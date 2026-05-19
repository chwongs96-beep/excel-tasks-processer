"""
Excel report generation from templates (openpyxl).

Loads a per-report-type .xlsx template, writes transformed tables into configured
sheets/ranges, preserves existing cell styles, and saves to the output directory.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from reporting.core.dates import format_period_display, period_bounds
from reporting.core.filenames import build_output_filename as _build_output_filename
from reporting.core.safe_io import (
    UnsafeOutputPathError,
    atomic_save_workbook,
    atomic_write_bytes,
    secure_output_path,
)
from reporting.export.template_bootstrap import ensure_templates_exist
from reporting.export.template_config import (
    TableMapping,
    TemplateConfig,
    TemplateSpec,
    load_template_config,
)
from reporting.export.workbook_builder import build_workbook_bytes
from reporting.models import ReportType, ValidationIssue

ReportTables = dict[str, pd.DataFrame]


@dataclass
class ExportResult:
    """File path and metadata returned to the UI / pipeline."""

    file_path: Path
    filename: str
    workbook_bytes: bytes
    metadata: dict[str, Any] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    issues: list[ValidationIssue] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Template loading
# ---------------------------------------------------------------------------


def resolve_template_path(report_type: ReportType, config: TemplateConfig | None = None) -> Path:
    """Return absolute path to the .xlsx template for a report type."""
    cfg = config or load_template_config()
    if report_type not in cfg.templates:
        raise KeyError(f"No template configured for report type: {report_type}")
    return cfg.template_path(report_type)


def load_template_workbook(
    report_type: ReportType,
    config: TemplateConfig | None = None,
) -> tuple[Workbook, TemplateSpec, Path]:
    """
    Load template workbook from disk; create default templates if missing.

    Returns:
        (workbook, template_spec, template_path)
    """
    cfg = config or load_template_config()
    ensure_templates_exist()
    path = resolve_template_path(report_type, cfg)
    if not path.is_file():
        raise FileNotFoundError(f"找不到報表範本：{path}")
    workbook = load_workbook(path)
    return workbook, cfg.templates[report_type], path


# ---------------------------------------------------------------------------
# Cell / range writing (values only — preserve template styles)
# ---------------------------------------------------------------------------


def write_metadata_cells(
    workbook: Workbook,
    spec: TemplateSpec,
    metadata: dict[str, Any],
) -> None:
    """Write report metadata into fixed template cells (first sheet)."""
    if not spec.metadata_cells:
        return
    worksheet = workbook.worksheets[0]
    for key, cell_ref in spec.metadata_cells.items():
        if key in metadata:
            worksheet[cell_ref] = metadata[key]


def write_dataframe_to_range(
    worksheet: Worksheet,
    frame: pd.DataFrame,
    mapping: TableMapping,
) -> int:
    """
    Write DataFrame values starting at ``mapping.data_start``.

    Returns:
        Number of data rows written (excluding header).
    """
    start_col_letter, start_row = coordinate_from_string(mapping.data_start)
    start_col = column_index_from_string(start_col_letter)
    export_frame = frame.map(lambda v: None if pd.isna(v) else v)

    row_offset = 0
    if mapping.write_header:
        for col_idx, column in enumerate(export_frame.columns, start=start_col):
            worksheet.cell(row=start_row, column=col_idx, value=column)
        row_offset = 1

    data_row_count = 0
    for row_idx, row in enumerate(export_frame.itertuples(index=False), start=0):
        target_row = start_row + row_offset + row_idx
        if mapping.style_reference_row:
            _copy_row_style(
                worksheet,
                mapping.style_reference_row,
                target_row,
                start_col,
                start_col + len(export_frame.columns) - 1,
            )
        for col_idx, value in enumerate(row, start=start_col):
            worksheet.cell(row=target_row, column=col_idx, value=value)
        data_row_count += 1

    return data_row_count


def _copy_row_style(
    worksheet: Worksheet,
    source_row: int,
    target_row: int,
    col_start: int,
    col_end: int,
) -> None:
    for col in range(col_start, col_end + 1):
        src = worksheet.cell(row=source_row, column=col)
        dest = worksheet.cell(row=target_row, column=col)
        if src.has_style:
            dest.font = copy(src.font)
            dest.fill = copy(src.fill)
            dest.border = copy(src.border)
            dest.alignment = copy(src.alignment)
            dest.number_format = src.number_format


def write_tables_to_template(
    workbook: Workbook,
    spec: TemplateSpec,
    tables: ReportTables,
    *,
    title_to_id: dict[str, str] | None = None,
) -> dict[str, int]:
    """
    Write all logical tables into template sheets.

    ``tables`` keys may be output ids or display titles; unknown keys are skipped.
    """
    title_to_id = title_to_id or {}
    row_counts: dict[str, int] = {}

    for table_key, frame in tables.items():
        output_id = title_to_id.get(table_key, table_key)
        mapping = spec.table_mappings.get(output_id)
        if mapping is None:
            continue
        if mapping.sheet not in workbook.sheetnames:
            raise ValueError(f"範本缺少工作表：{mapping.sheet}")
        worksheet = workbook[mapping.sheet]
        count = write_dataframe_to_range(worksheet, frame, mapping)
        row_counts[mapping.sheet] = count

    return row_counts


# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------


def build_output_filename(
    report_type: ReportType,
    date_spec: dict[str, Any],
    config: TemplateConfig | None = None,
) -> str:
    cfg = config or load_template_config()
    return _build_output_filename(
        report_type,
        date_spec,
        filename_template=cfg.export.filename_template,
    )


def ensure_output_directory(config: TemplateConfig | None = None) -> Path:
    cfg = config or load_template_config()
    output_dir = cfg.output_directory()
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def save_workbook_to_path(
    workbook: Workbook,
    file_path: Path,
    *,
    output_dir: Path | None = None,
) -> Path:
    """Save workbook to disk with path checks and atomic write."""
    if output_dir is not None:
        file_path = secure_output_path(output_dir, file_path.name)
    return atomic_save_workbook(workbook, file_path)


def workbook_to_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def build_metadata(
    report_type: ReportType,
    date_spec: dict[str, Any],
    template_path: Path,
) -> dict[str, Any]:
    period_start, period_end = period_bounds(report_type, date_spec)
    labels = {"daily": "日報", "weekly": "週報", "monthly": "月報"}
    return {
        "report_type": report_type,
        "report_title": labels.get(report_type, report_type),
        "period_label": format_period_display(
            report_type,
            date_spec,
            start=period_start,
            end=period_end,
        ),
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "template_file": template_path.name,
    }


def build_report_from_template(
    tables: ReportTables,
    report_type: ReportType,
    date_spec: dict[str, Any],
    *,
    title_to_id: dict[str, str] | None = None,
) -> ExportResult:
    """
    Generate a formatted Excel report from a template.

    Falls back to programmatic workbook creation if the template cannot be loaded.
    """
    cfg = load_template_config()
    warnings: list[str] = []
    issues: list[ValidationIssue] = []

    filename = build_output_filename(report_type, date_spec, cfg)
    output_dir = ensure_output_directory(cfg)
    try:
        file_path = secure_output_path(output_dir, filename)
    except UnsafeOutputPathError as exc:
        return ExportResult(
            file_path=output_dir / "invalid.xlsx",
            filename=filename,
            workbook_bytes=b"",
            warnings=warnings,
            issues=[
                ValidationIssue(
                    code="unsafe_output_path",
                    message=str(exc),
                )
            ],
        )

    try:
        workbook, spec, template_path = load_template_workbook(report_type, cfg)
        metadata = build_metadata(report_type, date_spec, template_path)
        write_metadata_cells(workbook, spec, metadata)
        row_counts = write_tables_to_template(
            workbook, spec, tables, title_to_id=title_to_id
        )
        save_workbook_to_path(workbook, file_path)
        # file_path already secured above
        workbook_bytes = workbook_to_bytes(workbook)
        metadata.update({
            "output_path": str(file_path),
            "sheets_written": list(row_counts.keys()),
            "row_counts": row_counts,
            "used_template": True,
            "used_fallback": False,
        })
        return ExportResult(
            file_path=file_path,
            filename=filename,
            workbook_bytes=workbook_bytes,
            metadata=metadata,
            warnings=warnings,
            issues=issues,
        )
    except (FileNotFoundError, KeyError, ValueError) as exc:
        warnings.append(f"範本產生失敗，改用程式建立報表：{exc}")
        issues.append(
            ValidationIssue(
                code="template_fallback",
                message=str(exc),
                details={"report_type": report_type},
            )
        )
        return _build_fallback_report(
            tables, report_type, date_spec, file_path, filename, warnings, issues
        )


def _build_fallback_report(
    tables: ReportTables,
    report_type: ReportType,
    date_spec: dict[str, Any],
    file_path: Path,
    filename: str,
    warnings: list[str],
    issues: list[ValidationIssue],
) -> ExportResult:
    workbook_bytes = build_workbook_bytes(tables)
    atomic_write_bytes(workbook_bytes, file_path)
    period_start, period_end = period_bounds(report_type, date_spec)
    metadata = {
        "report_type": report_type,
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "output_path": str(file_path),
        "used_template": False,
        "used_fallback": True,
    }
    return ExportResult(
        file_path=file_path,
        filename=filename,
        workbook_bytes=workbook_bytes,
        metadata=metadata,
        warnings=warnings,
        issues=issues,
    )


def resolve_title_to_id_map(report_type: ReportType) -> dict[str, str]:
    """Map display titles (from transformer) to report_definitions output ids."""
    import app_config

    outputs = app_config.get_report_outputs(report_type)
    return {o.title: o.id for o in outputs}


