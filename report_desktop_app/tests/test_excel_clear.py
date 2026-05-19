"""Clear range in workbook."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.core.range_spec import SourceRangeSpec
from app.services.excel_clear_service import ExcelClearService
from app.services.excel_reader import ExcelReaderService


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "clear_me.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "H1"
    ws["A2"] = 100
    ws["B2"] = 200
    wb.save(path)
    wb.close()
    return path


def test_clear_range_clears_values(sample_xlsx: Path) -> None:
    spec = SourceRangeSpec(sheet="Data", excel_range="A1:B2")
    result = ExcelClearService().clear_range(sample_xlsx, spec)
    assert result.success
    assert result.cells_cleared == 4

    frame = ExcelReaderService().load_sheet(sample_xlsx, range_spec=spec)
    assert frame.isna().all().all() or (frame.fillna("") == "").all().all()
