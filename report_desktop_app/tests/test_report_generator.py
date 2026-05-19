"""Tests for Excel report generation."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.core import config
from app.core.schemas import DateSpec, ReportJob
from app.services.report_generator import (
    bootstrap_templates,
    build_output_filename,
    export_tables_to_excel,
)


def test_bootstrap_creates_templates(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(config, "TEMPLATES_DIR", tmp_path / "templates")
    monkeypatch.setattr(
        config,
        "TEMPLATE_FILES",
        {
            "daily": tmp_path / "templates" / "daily_report_template.xlsx",
            "weekly": tmp_path / "templates" / "weekly_report_template.xlsx",
            "monthly": tmp_path / "templates" / "monthly_report_template.xlsx",
        },
    )
    created = bootstrap_templates()
    assert len(created) == 3
    assert all(p.is_file() for p in created)


def test_export_daily_report(tmp_path, monkeypatch) -> None:
    templates_dir = tmp_path / "templates"
    output_dir = tmp_path / "output"
    templates_dir.mkdir()
    output_dir.mkdir()
    monkeypatch.setattr(config, "TEMPLATES_DIR", templates_dir)
    monkeypatch.setattr(config, "OUTPUT_DIR", output_dir)
    daily_path = templates_dir / "daily_report_template.xlsx"
    monkeypatch.setattr(
        config,
        "TEMPLATE_FILES",
        {
            "daily": daily_path,
            "weekly": templates_dir / "weekly_report_template.xlsx",
            "monthly": templates_dir / "monthly_report_template.xlsx",
        },
    )
    bootstrap_templates()

    tables = {
        "日報-依帳戶彙總": pd.DataFrame({"account_id": ["A1"], "amount": [100.0]}),
        "日報-明細": pd.DataFrame(
            {
                "trade_date": [date(2024, 1, 15)],
                "account_id": ["A1"],
                "symbol": ["TST"],
                "amount": [100.0],
                "currency": ["TWD"],
            }
        ),
    }
    result = export_tables_to_excel(
        tables,
        "daily",
        {"date": date(2024, 1, 15)},
        output_dir=output_dir,
        template_path=daily_path,
    )
    assert result.success
    assert result.file_path is not None
    assert result.file_path.is_file()
    assert result.filename == build_output_filename("daily", {"date": date(2024, 1, 15)})

    wb = load_workbook(result.file_path)
    assert "Summary" in wb.sheetnames
    assert wb["Summary"]["A5"].value == "A1"


def test_generate_job_with_canonical(tmp_path, monkeypatch) -> None:
    templates_dir = tmp_path / "templates"
    output_dir = tmp_path / "output"
    monkeypatch.setattr(config, "TEMPLATES_DIR", templates_dir)
    monkeypatch.setattr(config, "OUTPUT_DIR", output_dir)
    monkeypatch.setattr(
        config,
        "TEMPLATE_FILES",
        {
            "daily": templates_dir / "daily_report_template.xlsx",
            "weekly": templates_dir / "weekly_report_template.xlsx",
            "monthly": templates_dir / "monthly_report_template.xlsx",
        },
    )
    bootstrap_templates()

    canonical = pd.DataFrame(
        {
            "trade_date": pd.to_datetime(["2024-01-15"]),
            "account_id": ["A1"],
            "symbol": ["X"],
            "description": [None],
            "debit": [None],
            "credit": [None],
            "amount": [50.0],
            "currency": ["TWD"],
        }
    )
    from app.services.report_generator import ReportGeneratorService

    job = ReportJob(
        files=[],
        mapping={},
        report_type="daily",
        date_spec=DateSpec(report_type="daily", trade_date=date(2024, 1, 15)),
        output_dir=output_dir,
    )
    outcome = ReportGeneratorService().generate(job, merged_canonical=canonical)
    assert outcome.success
    assert outcome.output_path is not None
