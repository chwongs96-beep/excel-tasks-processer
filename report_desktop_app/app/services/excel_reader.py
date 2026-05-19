"""Excel import and preview (sheet + optional range)."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.core import config
from app.core.logger import get_logger
from app.core.range_spec import SourceRangeSpec
from app.core.schemas import LoadedFile

logger = get_logger()


@dataclass(frozen=True)
class PathUploadAdapter:
    """Makes a file path look like a Streamlit upload for shared validators."""

    path: Path

    @property
    def name(self) -> str:
        return self.path.name

    @property
    def size(self) -> int:
        return self.path.stat().st_size


class ExcelReaderService:
    """Read workbooks from disk for the desktop UI."""

    def __init__(self, *, preview_row_limit: int | None = None) -> None:
        self._preview_limit = preview_row_limit or config.PREVIEW_ROW_LIMIT

    def list_sheet_names(self, path: Path) -> list[str]:
        path = Path(path)
        if path.suffix.lower() == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                return list(workbook.sheetnames)
            finally:
                workbook.close()
        return ["Sheet1"]

    def inspect(
        self,
        path: Path,
        *,
        range_spec: SourceRangeSpec | None = None,
    ) -> LoadedFile:
        path = Path(path)
        suffix = path.suffix.lower()
        if suffix not in config.ALLOWED_EXTENSIONS:
            raise ValueError(f"不支援的檔案格式：{path.name}（僅支援 .xlsx、.xls）")

        spec = range_spec or SourceRangeSpec.default()
        sheet_names = self.list_sheet_names(path)

        if suffix == ".xlsx":
            columns, row_count = self._inspect_xlsx(path, spec, sheet_names)
        else:
            columns, row_count = self._inspect_xls(path, spec)

        if row_count > config.MAX_ROWS_PER_FILE:
            raise ValueError(
                f"{path.name} 列數超過上限（{row_count:,} > {config.MAX_ROWS_PER_FILE:,}）。"
            )

        logger.info("Inspected %s: %s columns, %s rows", path.name, len(columns), row_count)
        return LoadedFile(
            path=path,
            columns=columns,
            row_count=row_count,
            sheet_names=sheet_names,
            source_range=spec,
        )

    def _resolve_sheet_name(self, sheet_names: list[str], spec: SourceRangeSpec) -> str:
        if spec.sheet and spec.sheet in sheet_names:
            return spec.sheet
        return sheet_names[0] if sheet_names else "Sheet1"

    def _inspect_xlsx(
        self,
        path: Path,
        spec: SourceRangeSpec,
        sheet_names: list[str],
    ) -> tuple[list[str], int]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet_name = self._resolve_sheet_name(sheet_names, spec)
            sheet = workbook[sheet_name]
            header_row, min_row, max_row, min_col, max_col = spec.resolved_bounds()
            header_cells = next(
                sheet.iter_rows(
                    min_row=header_row,
                    max_row=header_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                ),
                (),
            )
            columns = [str(c) if c is not None else "" for c in header_cells]
            data_start = max(header_row + 1, min_row)
            effective_max = min(max_row, sheet.max_row or max_row)
            row_count = max(effective_max - data_start + 1, 0)
            if spec.end_row is None and spec.excel_range is None:
                row_count = max((sheet.max_row or 0) - header_row, 0)
            return columns, row_count
        finally:
            workbook.close()

    def _inspect_xls(self, path: Path, spec: SourceRangeSpec) -> tuple[list[str], int]:
        frame = self.load_sheet(path, range_spec=spec)
        return [str(c) for c in frame.columns], len(frame)

    def load_preview(
        self,
        path: Path,
        *,
        range_spec: SourceRangeSpec | None = None,
    ) -> pd.DataFrame:
        return self.load_sheet(path, range_spec=range_spec).head(self._preview_limit)

    def load_sheet(
        self,
        path: Path,
        *,
        range_spec: SourceRangeSpec | None = None,
    ) -> pd.DataFrame:
        path = Path(path)
        suffix = path.suffix.lower()
        if suffix not in config.ALLOWED_EXTENSIONS:
            raise ValueError(f"不支援的檔案格式：{path.name}")

        spec = range_spec or SourceRangeSpec.default()
        if suffix == ".xlsx":
            frame = self._load_xlsx_range(path, spec)
        else:
            frame = self._load_xls_range(path, spec)

        if len(frame) > config.MAX_ROWS_PER_FILE:
            raise ValueError(f"{path.name} 列數超過允許上限。")
        return frame

    def _load_xlsx_range(self, path: Path, spec: SourceRangeSpec) -> pd.DataFrame:
        sheet_names = self.list_sheet_names(path)
        sheet_name = self._resolve_sheet_name(sheet_names, spec)
        header_row, min_row, max_row, min_col, max_col = spec.resolved_bounds()

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name]
            header_cells = next(
                sheet.iter_rows(
                    min_row=header_row,
                    max_row=header_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                ),
                (),
            )
            columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(header_cells, 1)]

            data_start = max(header_row + 1, min_row)
            effective_max = min(max_row, sheet.max_row or max_row)
            rows: list[tuple] = []
            for row in sheet.iter_rows(
                min_row=data_start,
                max_row=effective_max,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            ):
                rows.append(row)

            return pd.DataFrame(rows, columns=columns[: len(rows[0])] if rows else columns)
        finally:
            workbook.close()

    def _load_xls_range(self, path: Path, spec: SourceRangeSpec) -> pd.DataFrame:
        header_row, min_row, max_row, _, _ = spec.resolved_bounds()
        skip = header_row - 1
        nrows = None
        if spec.end_row is not None or spec.excel_range is not None:
            nrows = max_row - min_row + 1
        frame = pd.read_excel(path, engine="xlrd", sheet_name=0, skiprows=skip, nrows=nrows)
        return frame
