"""
Excel report generation using openpyxl templates.

Layers:
    1. Template bootstrap / load   — app/templates/*.xlsx
    2. Template write              — metadata + table ranges (preserve styles)
    3. Save                        — output/ with typed filenames

Driven by config/template_mapping.yaml (sheet names, cell addresses, output ids).
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet

from app.core import config
from app.core.dates import format_period_display, period_bounds
from app.core.reporting_bridge import ensure_reporting_package

ensure_reporting_package()

from reporting.core.filenames import build_output_filename as _build_output_filename  # noqa: E402
from reporting.core.safe_io import (  # noqa: E402
    UnsafeOutputPathError,
    atomic_save_workbook,
    atomic_write_bytes,
    secure_output_path,
)
from app.core.logger import get_logger
from app.core.schemas import DateSpec, ReportJob, ReportOutcome, ReportType, TableMapping, TemplateSpec
from app.services.transformer import TransformerService

ReportTables = dict[str, pd.DataFrame]
logger = get_logger()

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_ALIGN = Alignment(horizontal="center")


@dataclass
class ExportResult:
    """Internal result from template export (mapped to ReportOutcome for UI)."""

    success: bool
    file_path: Path | None = None
    filename: str | None = None
    metadata: dict[str, Any] = field(default_factory=dict)
    messages: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Template bootstrap
# ---------------------------------------------------------------------------


def _style_header_row(worksheet: Worksheet, row: int, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = worksheet.cell(row=row, column=col, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _add_report_sheet(workbook: Workbook, name: str, headers: list[str]) -> None:
    ws = workbook.create_sheet(name)
    ws["B2"] = "報表標題"
    ws["B3"] = "期間"
    ws["B4"] = "產生時間"
    _style_header_row(ws, 4, headers)


def _create_daily_template(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _add_report_sheet(wb, "Summary", ["account_id", "amount"])
    _add_report_sheet(wb, "Detail", ["trade_date", "account_id", "symbol", "amount", "currency"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _create_weekly_template(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _add_report_sheet(wb, "Summary", ["account_id", "amount"])
    _add_report_sheet(wb, "BySymbol", ["symbol", "amount"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _create_monthly_template(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _add_report_sheet(wb, "Summary", ["account_id", "amount", "debit", "credit"])
    _add_report_sheet(
        wb,
        "Detail",
        ["trade_date", "account_id", "symbol", "amount", "debit", "credit", "currency"],
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


_BOOTSTRAP_BY_TYPE = {
    "daily": _create_daily_template,
    "weekly": _create_weekly_template,
    "monthly": _create_monthly_template,
}


def bootstrap_templates() -> list[Path]:
    """Create missing templates under app/templates/. Returns newly created paths."""
    config.ensure_dirs()
    created: list[Path] = []
    for report_type, path in config.TEMPLATE_FILES.items():
        if path.is_file():
            continue
        creator = _BOOTSTRAP_BY_TYPE.get(report_type)
        if creator is None:
            raise FileNotFoundError(f"無法建立範本：{report_type}")
        creator(path)
        created.append(path)
    return created


# ---------------------------------------------------------------------------
# Template loading
# ---------------------------------------------------------------------------


def load_template_workbook(
    report_type: ReportType,
    *,
    template_path: Path | None = None,
) -> tuple[Workbook, TemplateSpec, Path]:
    """Load an existing .xlsx template workbook."""
    bootstrap_templates()
    template_cfg = config.load_template_config()
    if report_type not in template_cfg.templates:
        raise KeyError(f"未設定報表範本：{report_type}")

    path = template_path or config.resolve_template_path(report_type)
    if not path.is_file():
        raise FileNotFoundError(f"找不到報表範本：{path}")

    workbook = load_workbook(path)
    return workbook, template_cfg.templates[report_type], path


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------


def write_metadata_cells(
    workbook: Workbook,
    spec: TemplateSpec,
    metadata: dict[str, Any],
) -> None:
    """Write fixed metadata cells (typically on the first sheet)."""
    if not spec.metadata_cells or not workbook.worksheets:
        return
    worksheet = workbook.worksheets[0]
    for key, cell_ref in spec.metadata_cells.items():
        if key in metadata:
            worksheet[cell_ref] = metadata[key]


def write_dataframe_to_range(
    worksheet: Worksheet,
    frame: pd.DataFrame,
    mapping: TableMapping,
) -> int:
    """Write values starting at mapping.data_start; copy styles from reference row."""
    start_col_letter, start_row = coordinate_from_string(mapping.data_start)
    start_col = column_index_from_string(start_col_letter)
    export_frame = frame.map(lambda v: None if pd.isna(v) else v)

    row_offset = 0
    if mapping.write_header:
        for col_idx, column in enumerate(export_frame.columns, start=start_col):
            worksheet.cell(row=start_row, column=col_idx, value=column)
        row_offset = 1

    data_rows = 0
    for row_idx, row in enumerate(export_frame.itertuples(index=False), start=0):
        target_row = start_row + row_offset + row_idx
        if mapping.style_reference_row:
            _copy_row_style(
                worksheet,
                mapping.style_reference_row,
                target_row,
                start_col,
                start_col + len(export_frame.columns) - 1,
            )
        for col_idx, value in enumerate(row, start=start_col):
            worksheet.cell(row=target_row, column=col_idx, value=value)
        data_rows += 1
    return data_rows


def _copy_row_style(
    worksheet: Worksheet,
    source_row: int,
    target_row: int,
    col_start: int,
    col_end: int,
) -> None:
    for col in range(col_start, col_end + 1):
        src = worksheet.cell(row=source_row, column=col)
        dest = worksheet.cell(row=target_row, column=col)
        if src.has_style:
            dest.font = copy(src.font)
            dest.fill = copy(src.fill)
            dest.border = copy(src.border)
            dest.alignment = copy(src.alignment)
            dest.number_format = src.number_format


def write_tables_to_template(
    workbook: Workbook,
    spec: TemplateSpec,
    tables: ReportTables,
    *,
    title_to_id: dict[str, str] | None = None,
) -> dict[str, int]:
    """Write all tables; keys may be output id or display title."""
    title_to_id = title_to_id or {}
    row_counts: dict[str, int] = {}

    for table_key, frame in tables.items():
        output_id = title_to_id.get(table_key, table_key)
        mapping = spec.table_mappings.get(output_id)
        if mapping is None:
            continue
        if mapping.sheet not in workbook.sheetnames:
            raise ValueError(f"範本缺少工作表：{mapping.sheet}")
        worksheet = workbook[mapping.sheet]
        count = write_dataframe_to_range(worksheet, frame, mapping)
        row_counts[mapping.sheet] = count
    return row_counts


# ---------------------------------------------------------------------------
# Saving & filenames
# ---------------------------------------------------------------------------


def build_output_filename(report_type: ReportType, date_spec: dict[str, Any]) -> str:
    template_cfg = config.load_template_config()
    return _build_output_filename(
        report_type,
        date_spec,
        filename_template=template_cfg.filename_template,
    )


def build_metadata(
    report_type: ReportType,
    date_spec: dict[str, Any],
    template_path: Path,
) -> dict[str, Any]:
    period_start, period_end = period_bounds(report_type, date_spec)
    labels = {"daily": "日報", "weekly": "週報", "monthly": "月報"}
    return {
        "report_type": report_type,
        "report_title": labels.get(report_type, report_type),
        "period_label": format_period_display(
            report_type, date_spec, start=period_start, end=period_end
        ),
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "template_file": template_path.name,
    }


def save_workbook(workbook: Workbook, file_path: Path, *, output_dir: Path | None = None) -> Path:
    if output_dir is not None:
        file_path = secure_output_path(output_dir, file_path.name)
    return atomic_save_workbook(workbook, file_path)


def workbook_to_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Fallback workbook (no template)
# ---------------------------------------------------------------------------


def build_fallback_workbook(tables: ReportTables) -> bytes:
    """Create a simple multi-sheet workbook when template load/write fails."""
    wb = Workbook()
    wb.remove(wb.active)
    for title, frame in tables.items():
        safe_name = str(title)[:31]
        ws = wb.create_sheet(safe_name)
        for col_idx, column in enumerate(frame.columns, start=1):
            ws.cell(row=1, column=col_idx, value=column)
        for row_idx, row in enumerate(frame.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=None if pd.isna(value) else value,
                )
    return workbook_to_bytes(wb)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------


def export_tables_to_excel(
    tables: ReportTables,
    report_type: ReportType,
    date_spec: dict[str, Any],
    *,
    output_dir: Path | None = None,
    template_path: Path | None = None,
) -> ExportResult:
    """Load template, write tables, save to output directory."""
    if not tables:
        return ExportResult(
            success=False,
            errors=["沒有可寫入報表的資料表。"],
        )

    out_dir = Path(output_dir or config.OUTPUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = build_output_filename(report_type, date_spec)
    try:
        file_path = secure_output_path(out_dir, filename)
    except UnsafeOutputPathError as exc:
        return ExportResult(success=False, errors=[str(exc)])
    title_to_id = config.title_to_output_id_map(report_type)

    try:
        workbook, spec, resolved_template = load_template_workbook(
            report_type,
            template_path=template_path,
        )
        metadata = build_metadata(report_type, date_spec, resolved_template)
        write_metadata_cells(workbook, spec, metadata)
        row_counts = write_tables_to_template(
            workbook,
            spec,
            tables,
            title_to_id=title_to_id,
        )
        save_workbook(workbook, file_path)
        metadata.update({
            "output_path": str(file_path),
            "sheets_written": list(row_counts.keys()),
            "row_counts": row_counts,
            "used_template": True,
        })
        return ExportResult(
            success=True,
            file_path=file_path,
            filename=filename,
            metadata=metadata,
            messages=[f"報表已儲存：{file_path}"],
        )
    except (FileNotFoundError, KeyError, ValueError, OSError) as exc:
        logger.warning("Template export failed, using fallback: %s", exc)
        try:
            atomic_write_bytes(build_fallback_workbook(tables), file_path)
            return ExportResult(
                success=True,
                file_path=file_path,
                filename=filename,
                metadata={"used_template": False, "used_fallback": True},
                messages=[
                    "範本寫入失敗，已改用簡易 Excel 格式輸出。",
                    "若需正式版面，請檢查範本檔與 template_mapping.yaml。",
                    f"技術細節：{exc}",
                ],
            )
        except OSError as save_exc:
            return ExportResult(
                success=False,
                errors=[f"無法儲存報表：{save_exc}"],
            )


def _export_result_to_outcome(result: ExportResult, tables: ReportTables) -> ReportOutcome:
    return ReportOutcome(
        success=result.success,
        tables=tables,
        output_path=result.file_path,
        filename=result.filename,
        messages=result.messages,
        errors=result.errors,
        metadata=result.metadata,
    )


# ---------------------------------------------------------------------------
# Service facade
# ---------------------------------------------------------------------------


class ReportGeneratorService:
    """Generate Excel reports for the desktop application."""

    def __init__(self) -> None:
        self._transformer = TransformerService()

    def ensure_templates(self) -> list[Path]:
        return bootstrap_templates()

    def export(
        self,
        tables: ReportTables,
        report_type: ReportType,
        date_spec: DateSpec,
        *,
        output_dir: Path | None = None,
        template_path: Path | None = None,
    ) -> ReportOutcome:
        result = export_tables_to_excel(
            tables,
            report_type,
            date_spec.to_dict(),
            output_dir=output_dir,
            template_path=template_path,
        )
        return _export_result_to_outcome(result, tables)

    def generate(
        self,
        job: ReportJob,
        *,
        tables: ReportTables | None = None,
        merged_canonical: pd.DataFrame | None = None,
    ) -> ReportOutcome:
        """
        End-to-end: transform merged data (if needed) then export to Excel.

        Provide ``tables`` to skip transformation, or ``merged_canonical`` to run
        period filter + aggregation inside this service.
        """
        bootstrap_templates()

        if tables is None:
            if merged_canonical is None or merged_canonical.empty:
                return ReportOutcome(
                    success=False,
                    errors=["沒有可產報的資料，請先匯入並驗證 Excel。"],
                )
            transform = self._transformer.build_report_tables(
                merged_canonical,
                job.report_type,
                job.date_spec,
            )
            if transform.blocking:
                return ReportOutcome(
                    success=False,
                    errors=[transform.blocking_message or "轉換失敗。"],
                    messages=transform.warnings,
                )
            tables = transform.tables

        return self.export(
            tables,
            job.report_type,
            job.date_spec,
            output_dir=job.output_dir,
            template_path=job.template_path,
        )


