"""Merge selected ranges from multiple Excel files into one workbook."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Literal

import pandas as pd

from app.core.progress import ProgressReporter
from openpyxl import Workbook, load_workbook

from app.core.logger import get_logger
from app.core.range_spec import SourceRangeSpec
from app.core.reporting_bridge import ensure_reporting_package
from app.services.excel_reader import ExcelReaderService

ensure_reporting_package()

from reporting.core.safe_io import atomic_save_workbook, secure_output_path  # noqa: E402

logger = get_logger()

MergeMode = Literal["single_sheet", "one_sheet_per_file"]


@dataclass
class ConsolidateRequest:
    """User choices for multi-file merge."""

    sources: list[tuple[Path, SourceRangeSpec]]
    output_path: Path
    merge_mode: MergeMode = "single_sheet"
    use_template: bool = False
    template_path: Path | None = None
    combined_sheet_name: str = "合併資料"
    import_after_merge: bool = False
    open_mapping_after_merge: bool = False


@dataclass
class ConsolidateResult:
    success: bool
    output_path: Path | None = None
    messages: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    row_counts: dict[str, int] = field(default_factory=dict)


class ConsolidationService:
    """Copy ranged data from many workbooks into one output .xlsx."""

    def __init__(self) -> None:
        self._reader = ExcelReaderService()

    def consolidate(
        self,
        request: ConsolidateRequest,
        progress: ProgressReporter | None = None,
    ) -> ConsolidateResult:
        if not request.sources:
            return ConsolidateResult(success=False, errors=["請至少選擇一個來源檔案。"])

        if progress:
            progress.start(0, "檢查輸出路徑與檔名")
            progress.log(f"共 {len(request.sources)} 個來源檔")

        try:
            output_path = secure_output_path(
                request.output_path.parent,
                request.output_path.name,
            )
        except Exception as exc:  # noqa: BLE001
            if progress:
                progress.done(0)
            return ConsolidateResult(success=False, errors=[str(exc)])

        if progress:
            progress.done(0)

        frames: list[tuple[Path, pd.DataFrame]] = []
        for index, (path, range_spec) in enumerate(request.sources):
            step = 1 + index
            if progress:
                progress.start(step, f"讀取 {path.name}")
            try:
                frame = self._reader.load_sheet(path, range_spec=range_spec)
                if frame.empty:
                    logger.warning("Empty range: %s", path.name)
                frames.append((path, frame))
                if progress:
                    progress.log(f"  → {len(frame):,} 列")
            except Exception as exc:  # noqa: BLE001
                logger.exception("Failed to read %s", path)
                if progress:
                    progress.done(step)
                return ConsolidateResult(
                    success=False,
                    errors=[f"讀取「{path.name}」失敗：{exc}"],
                )
            if progress:
                progress.done(step)

        if not frames:
            return ConsolidateResult(success=False, errors=["沒有可合併的資料列。"])

        write_step = 1 + len(request.sources)
        if progress:
            progress.start(write_step, "寫入合併 Excel")
        try:
            if request.use_template and request.template_path:
                path, counts = self._write_with_template(request, frames, output_path)
            else:
                path, counts = self._write_plain_workbook(request, frames, output_path)
        except Exception as exc:  # noqa: BLE001
            logger.exception("Consolidate write failed")
            if progress:
                progress.done(write_step)
            return ConsolidateResult(success=False, errors=[f"寫入失敗：{exc}"])

        if progress:
            progress.log(
                "；".join(f"{name} {n:,} 列" for name, n in counts.items())
            )
            progress.done(write_step)

        return ConsolidateResult(
            success=True,
            output_path=path,
            messages=[f"已合併 {len(frames)} 個來源至：{path}"],
            row_counts=counts,
        )

    def _write_plain_workbook(
        self,
        request: ConsolidateRequest,
        frames: list[tuple[Path, pd.DataFrame]],
        output_path: Path,
    ) -> tuple[Path, dict[str, int]]:
        counts: dict[str, int] = {}
        wb = Workbook()
        wb.remove(wb.active)

        if request.merge_mode == "single_sheet":
            combined = pd.concat(
                [f.assign(_source_file=path.name) for path, f in frames],
                ignore_index=True,
            )
            ws = wb.create_sheet(request.combined_sheet_name[:31])
            self._write_frame_to_sheet(ws, combined)
            counts[request.combined_sheet_name] = len(combined)
        else:
            for path, frame in frames:
                safe = self._unique_sheet_name(wb, path.stem)
                ws = wb.create_sheet(safe)
                self._write_frame_to_sheet(ws, frame)
                counts[safe] = len(frame)

        atomic_save_workbook(wb, output_path)
        wb.close()
        return output_path, counts

    def _write_with_template(
        self,
        request: ConsolidateRequest,
        frames: list[tuple[Path, pd.DataFrame]],
        output_path: Path,
    ) -> tuple[Path, dict[str, int]]:
        template_path = Path(request.template_path)
        if not template_path.is_file():
            raise FileNotFoundError(f"找不到範本：{template_path}")

        wb = load_workbook(template_path)
        counts: dict[str, int] = {}

        if request.merge_mode == "single_sheet":
            combined = pd.concat(
                [f.assign(_source_file=path.name) for path, f in frames],
                ignore_index=True,
            )
            target = wb.worksheets[0]
            self._write_frame_to_sheet(target, combined, start_row=2)
            counts[target.title] = len(combined)
        else:
            for path, frame in frames:
                safe = self._unique_sheet_name(wb, path.stem)
                ws = wb.create_sheet(safe)
                self._write_frame_to_sheet(ws, frame)
                counts[safe] = len(frame)

        atomic_save_workbook(wb, output_path)
        wb.close()
        return output_path, counts

    @staticmethod
    def _write_frame_to_sheet(ws, frame: pd.DataFrame, *, start_row: int = 1) -> None:
        for col_idx, column in enumerate(frame.columns, start=1):
            ws.cell(row=start_row, column=col_idx, value=column)
        for row_offset, row in enumerate(frame.itertuples(index=False), start=1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(
                    row=start_row + row_offset,
                    column=col_idx,
                    value=None if pd.isna(value) else value,
                )

    @staticmethod
    def _unique_sheet_name(wb: Workbook, base: str) -> str:
        name = base[:31] or "Sheet"
        if name not in wb.sheetnames:
            return name
        for index in range(2, 100):
            candidate = f"{name[:28]}_{index}"
            if candidate not in wb.sheetnames:
                return candidate
        return "Sheet_extra"
