"""Clear cell contents in a selected Excel range."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.core.logger import get_logger
from app.core.range_spec import SourceRangeSpec, parse_excel_range
from app.core.reporting_bridge import ensure_reporting_package

ensure_reporting_package()

from reporting.core.safe_io import atomic_save_workbook  # noqa: E402

logger = get_logger()


@dataclass
class ClearRangeResult:
    success: bool
    cells_cleared: int = 0
    message: str = ""
    error: str = ""


class ExcelClearService:
    """Clear values (and optionally formulas) inside a workbook range."""

    def clear_range(
        self,
        path: Path,
        spec: SourceRangeSpec,
        *,
        contents_only: bool = True,
    ) -> ClearRangeResult:
        path = Path(path)
        if not path.is_file():
            return ClearRangeResult(success=False, error=f"找不到檔案：{path}")

        sheet_name = spec.sheet
        try:
            wb = load_workbook(path)
        except Exception as exc:  # noqa: BLE001
            return ClearRangeResult(success=False, error=f"無法開啟活頁簿：{exc}")

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif sheet_name:
            wb.close()
            return ClearRangeResult(success=False, error=f"找不到工作表：{sheet_name}")
        else:
            ws = wb.active

        try:
            min_row, max_row, min_col, max_col = self._resolve_clear_bounds(spec)
        except ValueError as exc:
            wb.close()
            return ClearRangeResult(success=False, error=str(exc))

        cleared = 0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.value = None
                if not contents_only and getattr(cell, "data_type", None) == "f":
                    cell.value = None
                cleared += 1

        try:
            atomic_save_workbook(wb, path)
        except Exception as exc:  # noqa: BLE001
            wb.close()
            return ClearRangeResult(success=False, error=f"儲存失敗：{exc}")
        finally:
            wb.close()

        summary = spec.summary()
        logger.info("Cleared %s cells in %s (%s)", cleared, path.name, summary)
        return ClearRangeResult(
            success=True,
            cells_cleared=cleared,
            message=f"已清除 {cleared:,} 個儲存格（{summary}）",
        )

    @staticmethod
    def _resolve_clear_bounds(spec: SourceRangeSpec) -> tuple[int, int, int, int]:
        """Return min_row, max_row, min_col, max_col (1-based inclusive)."""
        if spec.excel_range:
            parsed = parse_excel_range(spec.excel_range)
            if parsed is None:
                raise ValueError(f"無效的 Excel 範圍：{spec.excel_range}")
            min_col, min_row, max_col, max_row = parsed
            return min_row, max_row, min_col, max_col

        _, min_row, max_row, min_col, max_col = spec.resolved_bounds()
        if spec.header_row and spec.start_row is None and not spec.excel_range:
            # If only header row set without data rows, clear from header row
            min_row = max(1, spec.header_row)
        return min_row, max_row, min_col, max_col
