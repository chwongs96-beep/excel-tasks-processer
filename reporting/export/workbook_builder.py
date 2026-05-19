"""Build formatted Excel workbooks from report tables."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook

from reporting.export.sheet_writers import write_dataframe_sheet


def build_workbook_bytes(tables: dict[str, pd.DataFrame]) -> bytes:
    """
    Create an in-memory .xlsx workbook from named tables.

    Each key becomes a worksheet title (truncated to 31 chars for Excel).
    """
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, frame in tables.items():
        safe_name = _safe_sheet_name(sheet_name)
        worksheet = workbook.create_sheet(title=safe_name)
        write_dataframe_sheet(worksheet, frame)

    if not workbook.sheetnames:
        workbook.create_sheet("empty")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _safe_sheet_name(name: str) -> str:
    invalid = set(r'[]:*?/\\')
    cleaned = "".join("_" if ch in invalid else ch for ch in name).strip()
    if not cleaned:
        cleaned = "sheet"
    return cleaned[:31]
