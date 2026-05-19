"""Write pandas DataFrames to openpyxl worksheets with formatting."""

from __future__ import annotations

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from reporting.export.styles import (
    AMOUNT_COLUMNS,
    AMOUNT_FORMAT,
    DATE_COLUMNS,
    DATE_FORMAT,
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
)


def write_dataframe_sheet(worksheet: Worksheet, frame: pd.DataFrame) -> None:
    """Write DataFrame to sheet and apply header/number formats."""
    export_frame = frame.copy()
    export_frame = export_frame.map(lambda value: None if pd.isna(value) else value)
    for row_idx, row in enumerate(
        dataframe_to_rows(export_frame, index=False, header=True), start=1
    ):
        worksheet.append(row)
        if row_idx == 1:
            for cell in worksheet[row_idx]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT

    _apply_column_formats(worksheet, list(frame.columns))
    worksheet.freeze_panes = "A2"
    _autosize_columns(worksheet)


def _apply_column_formats(worksheet: Worksheet, columns: list[str]) -> None:
    for col_idx, name in enumerate(columns, start=1):
        letter = worksheet.cell(row=1, column=col_idx).column_letter
        if name in DATE_COLUMNS:
            for row in range(2, worksheet.max_row + 1):
                worksheet[f"{letter}{row}"].number_format = DATE_FORMAT
        if name in AMOUNT_COLUMNS:
            for row in range(2, worksheet.max_row + 1):
                worksheet[f"{letter}{row}"].number_format = AMOUNT_FORMAT


def _autosize_columns(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 40)
