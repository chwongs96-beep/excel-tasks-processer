"""Create default Excel templates (run once or when templates are missing)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from reporting.export.template_config import PROJECT_ROOT, load_template_config

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_ALIGN = Alignment(horizontal="center")


def _style_header_row(worksheet, row: int, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = worksheet.cell(row=row, column=col, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _add_report_sheet(workbook: Workbook, name: str, headers: list[str]) -> None:
    ws = workbook.create_sheet(name)
    ws["B2"] = "報表標題"
    ws["B3"] = "期間"
    ws["B4"] = "產生時間"
    _style_header_row(ws, 4, headers)


def create_daily_template(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _add_report_sheet(wb, "Summary", ["account_id", "amount"])
    _add_report_sheet(wb, "Detail", ["trade_date", "account_id", "symbol", "amount", "currency"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def create_weekly_template(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _add_report_sheet(wb, "Summary", ["account_id", "amount"])
    _add_report_sheet(wb, "BySymbol", ["symbol", "amount"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def create_monthly_template(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _add_report_sheet(wb, "Summary", ["account_id", "amount", "debit", "credit"])
    _add_report_sheet(
        wb,
        "Detail",
        ["trade_date", "account_id", "symbol", "amount", "debit", "credit", "currency"],
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


_CREATORS = {
    "daily_report.xlsx": create_daily_template,
    "weekly_report.xlsx": create_weekly_template,
    "monthly_report.xlsx": create_monthly_template,
}


def ensure_templates_exist(config_path: Path | None = None) -> list[Path]:
    """Create any missing template files defined in template_mapping.yaml."""
    cfg = load_template_config(config_path)
    created: list[Path] = []
    for spec in cfg.templates.values():
        path = cfg.project_root / cfg.export.templates_dir / spec.file
        if path.exists():
            continue
        creator = _CREATORS.get(spec.file)
        if creator is None:
            raise FileNotFoundError(f"No bootstrap creator for template: {spec.file}")
        creator(path)
        created.append(path)
    return created
