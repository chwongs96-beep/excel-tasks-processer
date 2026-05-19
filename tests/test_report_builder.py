"""Tests for template-based report_builder."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

import report_builder as rb
from reporting.export.template_bootstrap import ensure_templates_exist
from reporting.export.template_config import load_template_config


@pytest.fixture(scope="module", autouse=True)
def _templates_ready() -> None:
    ensure_templates_exist()


def test_load_template_workbook_daily() -> None:
    wb, spec, path = rb.load_template_workbook("daily")
    assert path.name == spec.file
    assert "Summary" in wb.sheetnames
    wb.close()


def test_build_output_filename() -> None:
    name = rb.build_output_filename("daily", {"date": date(2026, 5, 1)})
    assert name.startswith("daily_")
    assert name.endswith(".xlsx")


def test_build_report_from_template(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(rb, "ensure_output_directory", lambda config=None: tmp_path)

    tables = {
        "日報-依帳戶彙總": pd.DataFrame({"account_id": ["A1"], "amount": [100]}),
        "日報-明細": pd.DataFrame({
            "trade_date": pd.to_datetime(["2026-05-01"]),
            "account_id": ["A1"],
            "symbol": [pd.NA],
            "amount": [100],
            "currency": ["TWD"],
        }),
    }
    title_map = rb.resolve_title_to_id_map("daily")
    result = rb.build_report_from_template(
        tables,
        "daily",
        {"date": date(2026, 5, 1)},
        title_to_id=title_map,
    )

    assert result.file_path.exists()
    assert result.metadata["used_template"] is True
    assert "Summary" in result.metadata["sheets_written"]

    wb = load_workbook(BytesIO(result.workbook_bytes))
    assert wb["Summary"]["A5"].value == "A1"
    wb.close()


def test_missing_template_falls_back(tmp_path, monkeypatch) -> None:
    cfg = load_template_config()
    monkeypatch.setattr(
        type(cfg),
        "output_directory",
        lambda self: tmp_path,
    )

    bad_path = cfg.project_root / cfg.export.templates_dir / "nonexistent.xlsx"

    def _fail_load(report_type, config=None):
        raise FileNotFoundError(f"找不到報表範本：{bad_path}")

    monkeypatch.setattr(rb, "load_template_workbook", _fail_load)

    tables = {"summary": pd.DataFrame({"a": [1]})}
    result = rb.build_report_from_template(
        tables,
        "daily",
        {"date": date(2026, 5, 1)},
    )
    assert result.metadata["used_fallback"] is True
    assert result.file_path.exists()
