"""Tests for Excel export."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from reporting.export.workbook_builder import build_workbook_bytes


def test_workbook_bytes_structure() -> None:
    tables = {
        "summary": pd.DataFrame({"account_id": ["A1"], "amount": [100]}),
        "detail": pd.DataFrame({
            "trade_date": pd.to_datetime(["2026-05-01"]),
            "amount": [100],
        }),
    }
    data = build_workbook_bytes(tables)
    wb = load_workbook(BytesIO(data))
    assert "summary" in wb.sheetnames
    assert "detail" in wb.sheetnames
    ws = wb["summary"]
    assert ws["A1"].value == "account_id"
    assert ws["B2"].value == 100
